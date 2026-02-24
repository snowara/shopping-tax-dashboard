"""
부가세 셀프 체크 도구
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
이카운트 매출/매입 엑셀 + 홈택스 세금계산서 목록을 비교하여
누락/불일치를 자동으로 찾아주는 스크립트

사용법:
    python3 vat_checker.py --quarter 2026Q1

입력 파일 (input/ 폴더에 넣기):
    - ecount_매출.xlsx  (이카운트 매출장 내보내기)
    - ecount_매입.xlsx  (이카운트 매입장 내보내기)
    - hometax_매출.xlsx (홈택스 전자세금계산서 매출 목록)
    - hometax_매입.xlsx (홈택스 전자세금계산서 매입 목록)

출력:
    - output/부가세체크_{quarter}.xlsx (대조 결과)
    - output/부가세체크_{quarter}.pdf  (요약 리포트)
"""
import os
import sys
import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = Path(__file__).parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 1. 엑셀 파일 자동 감지 및 로딩
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def find_file(keyword):
    """input 폴더에서 키워드가 포함된 엑셀 파일 찾기"""
    for f in INPUT_DIR.glob("*.xlsx"):
        if keyword in f.name.lower() or keyword in f.stem.lower():
            return f
    for f in INPUT_DIR.glob("*.xls"):
        if keyword in f.name.lower() or keyword in f.stem.lower():
            return f
    return None


def load_excel(filepath):
    """엑셀 파일 로딩 — 컬럼명 자동 감지"""
    if filepath is None:
        return None
    try:
        df = pd.read_excel(filepath)
        # 빈 행 제거
        df = df.dropna(how='all')
        return df
    except Exception as e:
        print(f"  ⚠️ 파일 로딩 실패: {filepath} → {e}")
        return None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 2. 컬럼 자동 매핑 (이카운트/홈택스 엑셀 형식 대응)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
COLUMN_MAP = {
    "date": ["일자", "날짜", "작성일자", "전송일자", "발급일자", "거래일자"],
    "partner": ["거래처", "거래처명", "공급받는자", "공급자", "상호"],
    "biz_no": ["사업자번호", "사업자등록번호", "등록번호"],
    "supply": ["공급가액", "공급가", "공급가합계", "금액"],
    "tax": ["세액", "부가세", "세액합계", "VAT"],
    "total": ["합계", "합계금액", "총액"],
    "item": ["품목", "품명", "품목명", "적요"],
    "slip_no": ["전표번호", "승인번호", "문서번호"],
}


def map_columns(df):
    """DataFrame 컬럼을 표준 이름으로 매핑"""
    mapped = {}
    for standard, candidates in COLUMN_MAP.items():
        for col in df.columns:
            col_clean = str(col).strip().replace(" ", "")
            for cand in candidates:
                if cand in col_clean:
                    mapped[standard] = col
                    break
            if standard in mapped:
                break
    return mapped


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 3. 대조 로직
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def normalize_amount(val):
    """금액 정규화"""
    if pd.isna(val):
        return 0
    if isinstance(val, str):
        val = val.replace(",", "").replace("원", "").replace(" ", "")
        try:
            return int(float(val))
        except ValueError:
            return 0
    return int(val)


def compare_data(ecount_df, hometax_df, label="매출"):
    """이카운트 vs 홈택스 데이터 대조"""
    results = {
        "label": label,
        "ecount_count": 0,
        "hometax_count": 0,
        "ecount_total": 0,
        "hometax_total": 0,
        "diff": 0,
        "missing_in_hometax": [],  # 이카운트에만 있음 (홈택스 누락)
        "missing_in_ecount": [],   # 홈택스에만 있음 (이카운트 누락)
        "amount_mismatch": [],     # 금액 불일치
    }

    if ecount_df is None and hometax_df is None:
        return results

    # 컬럼 매핑
    ec_map = map_columns(ecount_df) if ecount_df is not None else {}
    ht_map = map_columns(hometax_df) if hometax_df is not None else {}

    # 이카운트 합계
    if ecount_df is not None:
        results["ecount_count"] = len(ecount_df)
        if "supply" in ec_map:
            results["ecount_total"] = ecount_df[ec_map["supply"]].apply(normalize_amount).sum()
        elif "total" in ec_map:
            results["ecount_total"] = ecount_df[ec_map["total"]].apply(normalize_amount).sum()

    # 홈택스 합계
    if hometax_df is not None:
        results["hometax_count"] = len(hometax_df)
        if "supply" in ht_map:
            results["hometax_total"] = hometax_df[ht_map["supply"]].apply(normalize_amount).sum()
        elif "total" in ht_map:
            results["hometax_total"] = hometax_df[ht_map["total"]].apply(normalize_amount).sum()

    results["diff"] = results["ecount_total"] - results["hometax_total"]

    # 거래처 + 금액 기준 대조 (사업자번호가 있으면 사업자번호 우선)
    if ecount_df is not None and hometax_df is not None:
        # 매칭 키 결정
        ec_key = ec_map.get("biz_no") or ec_map.get("partner")
        ht_key = ht_map.get("biz_no") or ht_map.get("partner")
        ec_amt = ec_map.get("supply") or ec_map.get("total")
        ht_amt = ht_map.get("supply") or ht_map.get("total")

        if ec_key and ht_key and ec_amt and ht_amt:
            # 거래처별 합계 비교
            ec_grouped = ecount_df.groupby(ec_key)[ec_amt].apply(
                lambda x: x.apply(normalize_amount).sum()
            ).to_dict()
            ht_grouped = hometax_df.groupby(ht_key)[ht_amt].apply(
                lambda x: x.apply(normalize_amount).sum()
            ).to_dict()

            all_keys = set(list(ec_grouped.keys()) + list(ht_grouped.keys()))
            for key in all_keys:
                ec_val = ec_grouped.get(key, 0)
                ht_val = ht_grouped.get(key, 0)
                if ec_val > 0 and ht_val == 0:
                    results["missing_in_hometax"].append({
                        "거래처": key, "이카운트금액": ec_val
                    })
                elif ec_val == 0 and ht_val > 0:
                    results["missing_in_ecount"].append({
                        "거래처": key, "홈택스금액": ht_val
                    })
                elif abs(ec_val - ht_val) > 1:  # 1원 이상 차이
                    results["amount_mismatch"].append({
                        "거래처": key, "이카운트": ec_val, "홈택스": ht_val,
                        "차이": ec_val - ht_val
                    })

    return results


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 4. 리포트 생성
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
NAVY = "1B2A4A"
GOLD = "E8B931"
LIGHT_GRAY = "F5F5F5"


def create_report(sell_results, buy_results, quarter):
    """대조 결과를 엑셀 리포트로 생성"""
    wb = Workbook()

    # 스타일 정의
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
    warn_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    ok_fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ── 요약 시트 ──
    ws = wb.active
    ws.title = "요약"
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    from config import load_config
    cfg = load_config()
    ws.append([f"{cfg['company_name']} 부가세 셀프 체크", "", "", ""])
    ws.merge_cells("A1:D1")
    ws['A1'].font = Font(name="맑은 고딕", bold=True, size=16, color=NAVY)

    ws.append([f"기간: {quarter}", "", "작성일:", datetime.now().strftime("%Y-%m-%d")])
    ws.append([])

    # 요약 테이블
    headers = ["구분", "이카운트", "홈택스", "차이"]
    ws.append(headers)
    for col in range(1, 5):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for res in [sell_results, buy_results]:
        row = [
            f"{res['label']} 건수",
            f"{res['ecount_count']:,}건",
            f"{res['hometax_count']:,}건",
            f"{res['ecount_count'] - res['hometax_count']:+,}건"
        ]
        ws.append(row)
        row = [
            f"{res['label']} 공급가액",
            f"{res['ecount_total']:,.0f}원",
            f"{res['hometax_total']:,.0f}원",
            f"{res['diff']:+,.0f}원"
        ]
        ws.append(row)
        # 차이에 색상 적용
        r = ws.max_row
        diff_cell = ws.cell(row=r, column=4)
        if abs(res['diff']) > 0:
            diff_cell.fill = warn_fill
        else:
            diff_cell.fill = ok_fill

    ws.append([])

    # 부가세 예상 계산
    sell_tax = sell_results["ecount_total"] * 0.1
    buy_tax = buy_results["ecount_total"] * 0.1
    payable = sell_tax - buy_tax

    ws.append(["부가세 예상 (이카운트 기준)"])
    ws.cell(row=ws.max_row, column=1).font = Font(name="맑은 고딕", bold=True, size=12)
    ws.append(["매출세액 (10%)", f"{sell_tax:,.0f}원"])
    ws.append(["매입세액 (10%)", f"{buy_tax:,.0f}원"])
    ws.append(["납부예상세액", f"{payable:,.0f}원"])
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True, size=12, color=NAVY)

    # ── 누락/불일치 시트 ──
    for res in [sell_results, buy_results]:
        ws2 = wb.create_sheet(f"{res['label']}_불일치")

        if res["missing_in_hometax"]:
            ws2.append([f"홈택스 누락 (이카운트에만 있음) — {len(res['missing_in_hometax'])}건"])
            ws2.append(["거래처", "이카운트 금액"])
            for item in res["missing_in_hometax"]:
                ws2.append([item["거래처"], f"{item['이카운트금액']:,.0f}"])

        ws2.append([])

        if res["missing_in_ecount"]:
            ws2.append([f"이카운트 누락 (홈택스에만 있음) — {len(res['missing_in_ecount'])}건"])
            ws2.append(["거래처", "홈택스 금액"])
            for item in res["missing_in_ecount"]:
                ws2.append([item["거래처"], f"{item['홈택스금액']:,.0f}"])

        ws2.append([])

        if res["amount_mismatch"]:
            ws2.append([f"금액 불일치 — {len(res['amount_mismatch'])}건"])
            ws2.append(["거래처", "이카운트", "홈택스", "차이"])
            for item in res["amount_mismatch"]:
                ws2.append([
                    item["거래처"],
                    f"{item['이카운트']:,.0f}",
                    f"{item['홈택스']:,.0f}",
                    f"{item['차이']:+,.0f}"
                ])

        if not res["missing_in_hometax"] and not res["missing_in_ecount"] and not res["amount_mismatch"]:
            ws2.append(["✅ 불일치 항목 없음"])

    # 저장
    output_path = OUTPUT_DIR / f"부가세체크_{quarter}.xlsx"
    wb.save(output_path)
    return output_path


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 5. 메인 실행
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    parser = argparse.ArgumentParser(description="부가세 셀프 체크 도구")
    parser.add_argument("--quarter", default=None, help="분기 (예: 2026Q1)")
    args = parser.parse_args()

    # 분기 자동 설정
    now = datetime.now()
    q = (now.month - 1) // 3 + 1
    quarter = args.quarter or f"{now.year}Q{q}"

    from config import load_config
    cfg = load_config()

    print("=" * 50)
    print(f"  {cfg['company_name']} 부가세 셀프 체크")
    print(f"  기간: {quarter}")
    print("=" * 50)

    # 파일 탐색
    print("\n📂 input/ 폴더 파일 탐색...")
    files = {
        "ecount_매출": find_file("ecount") and find_file("매출"),
        "ecount_매입": find_file("ecount") and find_file("매입"),
        "hometax_매출": find_file("hometax") and find_file("매출"),
        "hometax_매입": find_file("hometax") and find_file("매입"),
    }

    # 좀 더 유연한 탐색
    ec_sell = find_file("매출") if not files["ecount_매출"] else files["ecount_매출"]
    ec_buy = find_file("매입") if not files["ecount_매입"] else files["ecount_매입"]

    all_files = list(INPUT_DIR.glob("*.xlsx")) + list(INPUT_DIR.glob("*.xls"))
    if not all_files:
        print("\n⚠️  input/ 폴더에 엑셀 파일이 없습니다.")
        print("\n📋 아래 파일을 넣어주세요:")
        print("   1. 이카운트 → 영업관리 → 매출장 → 엑셀 다운로드 → 'ecount_매출.xlsx'")
        print("   2. 이카운트 → 영업관리 → 매입장 → 엑셀 다운로드 → 'ecount_매입.xlsx'")
        print("   3. 홈택스 → 전자세금계산서 → 매출 목록조회 → 엑셀 다운 → 'hometax_매출.xlsx'")
        print("   4. 홈택스 → 전자세금계산서 → 매입 목록조회 → 엑셀 다운 → 'hometax_매입.xlsx'")
        print(f"\n   경로: {INPUT_DIR}/")
        return

    print(f"   발견된 파일: {len(all_files)}개")
    for f in all_files:
        print(f"   - {f.name}")

    # 데이터 로딩
    print("\n📊 데이터 로딩...")
    ec_sell_df = load_excel(ec_sell)
    ec_buy_df = load_excel(ec_buy)

    ht_sell = None
    ht_buy = None
    for f in all_files:
        name = f.name.lower()
        if "hometax" in name or "홈택스" in name:
            if "매출" in name or "sell" in name:
                ht_sell = f
            elif "매입" in name or "buy" in name:
                ht_buy = f

    ht_sell_df = load_excel(ht_sell)
    ht_buy_df = load_excel(ht_buy)

    # 대조 실행
    print("\n🔍 대조 실행...")
    sell_results = compare_data(ec_sell_df, ht_sell_df, "매출")
    buy_results = compare_data(ec_buy_df, ht_buy_df, "매입")

    # 결과 출력
    print(f"\n{'─'*50}")
    for res in [sell_results, buy_results]:
        print(f"\n  [{res['label']}]")
        print(f"  이카운트: {res['ecount_count']:,}건 / {res['ecount_total']:,.0f}원")
        print(f"  홈택스:   {res['hometax_count']:,}건 / {res['hometax_total']:,.0f}원")
        print(f"  차이:     {res['diff']:+,.0f}원")

        issues = len(res["missing_in_hometax"]) + len(res["missing_in_ecount"]) + len(res["amount_mismatch"])
        if issues > 0:
            print(f"  ⚠️  확인 필요: {issues}건")
        else:
            print(f"  ✅ 일치")

    # 리포트 생성
    print(f"\n{'─'*50}")
    output_path = create_report(sell_results, buy_results, quarter)
    print(f"\n📄 리포트 생성 완료: {output_path}")
    print(f"   open \"{output_path}\"")


if __name__ == "__main__":
    main()
