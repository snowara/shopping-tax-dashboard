"""
쇼핑몰 부가세 자료 수집 체크리스트 + 세무사 전달 패키징
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
8개 플랫폼 부가세 자료를 다운로드 → 정리 → 세무사 전달

사용법:
    python3 tax_package.py --quarter 2026Q1       (부가세)
    python3 tax_package.py --type corp --year 2025 (법인세)
"""
import argparse
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from config import load_config
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
INPUT_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

NAVY = "1B2A4A"
GOLD = "E8B931"
WHITE = "FFFFFF"
LIGHT_NAVY = "E8EDF5"
LIGHT_RED = "FFE8E8"
LIGHT_GREEN = "E8FFE8"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 8개 쇼핑몰 부가세 자료 정보
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PLATFORMS = [
    {
        "name": "스마트스토어",
        "seller_url": "https://sell.smartstore.naver.com/",
        "menu": "정산관리 > 부가세 신고 내역",
        "download": "기간 선택 → '월별내역 다운' 클릭",
        "file_format": "엑셀(.xlsx)",
        "filename": "스마트스토어",
    },
    {
        "name": "쿠팡",
        "seller_url": "https://wing.coupang.com/",
        "menu": "정산 > 부가세 신고 내역",
        "download": "기간 선택 → 상세 내역 다운로드",
        "file_format": "엑셀(.xlsx)",
        "filename": "쿠팡",
    },
    {
        "name": "11번가",
        "seller_url": "https://soffice.11st.co.kr/",
        "menu": "정산관리 > 부가세 신고 내역",
        "download": "월별 조회 → 인쇄(PDF) 또는 엑셀 다운",
        "file_format": "PDF/엑셀",
        "filename": "11번가",
    },
    {
        "name": "톡스토어",
        "seller_url": "https://business.kakao.com/",
        "menu": "정산관리 > 부가세 신고자료",
        "download": "기간 설정 → 엑셀 다운로드",
        "file_format": "엑셀(.xlsx)",
        "filename": "톡스토어",
    },
    {
        "name": "지그재그",
        "seller_url": "https://partner.kakaostyle.com/",
        "menu": "정산관리 > 국내 부가세 참고자료",
        "download": "기간 설정 → 엑셀 다운로드",
        "file_format": "엑셀(.xlsx)",
        "filename": "지그재그",
    },
    {
        "name": "롯데온",
        "seller_url": "https://partner.lotteon.com/",
        "menu": "정산관리 > 부가세 신고자료 조회",
        "download": "기간 선택 → 우측 상단 엑셀 아이콘 클릭",
        "file_format": "엑셀(.xlsx)",
        "filename": "롯데온",
    },
    {
        "name": "토스쇼핑",
        "seller_url": "https://shopping-seller.toss.im/",
        "menu": "쇼핑 > 정산내역",
        "download": "기간 선택 → 다운로드",
        "file_format": "확인 필요",
        "filename": "토스쇼핑",
    },
    {
        "name": "올웨이즈",
        "seller_url": "https://seller.alwayz.co/",
        "menu": "정산 > 세금계산서 조회",
        "download": "1기(1~6월)/2기(7~12월) 기간 선택 → 다운로드",
        "file_format": "확인 필요",
        "filename": "올웨이즈",
    },
]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 법인세 체크리스트
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CORP_CHECKLIST = [
    ("이카운트 재무제표", [
        ("재무상태표", "이카운트 자동 생성", "세무사 전달"),
        ("손익계산서", "이카운트 자동 생성", "세무사 전달"),
        ("이익잉여금처분계산서", "이카운트 자동 생성", "세무사 전달"),
        ("제조원가명세서", "이카운트 자동 생성", "세무사 전달"),
    ]),
    ("대표 준비 자료", [
        ("연간 매출장 (1~12월)", "이카운트 내보내기", "⚠️ 대표"),
        ("연간 매입장 (1~12월)", "이카운트 내보내기", "⚠️ 대표"),
        ("법인 통장 거래내역", "은행 내보내기", "⚠️ 대표"),
        ("임원 보수 지급 현황", "수동", "⚠️ 대표"),
    ]),
    ("대표 확인 항목", [
        ("감가상각 대상 자산", "해당시", "대표 확인"),
        ("가수금/가지급금 정리", "이카운트", "대표 확인"),
        ("접대비 연간 합계 (한도 3,600만원)", "이카운트", "대표 확인"),
        ("전기 대비 주요 변동사항", "수동", "⚠️ 대표 메모"),
    ]),
    ("기한", [
        ("법인세 신고: 사업연도 종료 후 3개월", "3월 31일", "필수"),
        ("세무조정료 결제", "세무사 청구", "대표"),
    ]),
]


def get_quarter_dates(quarter):
    """분기 문자열에서 시작/종료 날짜 반환"""
    year = int(quarter[:4])
    q = int(quarter[-1])
    start_month = (q - 1) * 3 + 1
    end_month = q * 3
    return f"{year}.{start_month:02d}.01", f"{year}.{end_month:02d}.{'30' if end_month in [4,6,9,11] else '31' if end_month in [1,3,5,7,8,10,12] else '28'}"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 스타일 헬퍼
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="맑은 고딕", bold=True, color=WHITE, size=10)
        cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )


def style_cell(cell, bold=False, color=None, fill=None, align="left"):
    cell.font = Font(name="맑은 고딕", bold=bold, size=10, color=color or "333333")
    if fill:
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 부가세 체크리스트 생성
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def create_vat_checklist(quarter):
    """8개 쇼핑몰 부가세 자료 수집 체크리스트"""
    wb = Workbook()
    ws = wb.active
    ws.title = "부가세 자료수집"

    # 컬럼 너비
    widths = [4, 14, 35, 28, 20, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    start_date, end_date = get_quarter_dates(quarter)

    # 타이틀
    ws.merge_cells("A1:F1")
    cfg = load_config()
    ws['A1'] = f"{cfg['company_name']} 부가세 자료 수집 체크리스트"
    ws['A1'].font = Font(name="맑은 고딕", bold=True, size=14, color=NAVY)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws['A2'] = f"기간: {quarter} ({start_date} ~ {end_date})  |  작성일: {datetime.now().strftime('%Y-%m-%d')}  |  세무사: {cfg['accountant_name']}"
    ws['A2'].font = Font(name="맑은 고딕", size=9, color="666666")

    # 헤더
    row = 4
    headers = ["#", "플랫폼", "메뉴 경로", "다운로드 방법", "파일 형식", "완료"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header(ws, row, len(headers))

    # 플랫폼별 행
    for i, p in enumerate(PLATFORMS, 1):
        row += 1
        ws.row_dimensions[row].height = 35

        values = [i, p["name"], p["menu"], p["download"], p["file_format"], "☐"]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_cell(cell, align="center" if col in [1, 5, 6] else "left")

        # 플랫폼명 강조
        ws.cell(row=row, column=2).font = Font(name="맑은 고딕", bold=True, size=10)

    # 파일 정리 안내
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="📁 다운로드 후 파일 정리")
    ws.cell(row=row, column=1).font = Font(name="맑은 고딕", bold=True, size=11, color=NAVY)

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value=f"다운로드한 파일을 아래 폴더에 넣어주세요: management-team/tax-automation/input/{quarter}/")
    ws.cell(row=row, column=1).font = Font(name="맑은 고딕", size=9, color="666666")

    file_rules = [
        ("스마트스토어_부가세.xlsx", "스마트스토어에서 다운로드한 파일"),
        ("쿠팡_부가세.xlsx", "쿠팡에서 다운로드한 파일"),
        ("11번가_부가세.xlsx (또는 .pdf)", "11번가에서 다운로드한 파일"),
        ("톡스토어_부가세.xlsx", "톡스토어에서 다운로드한 파일"),
        ("지그재그_부가세.xlsx", "지그재그에서 다운로드한 파일"),
        ("롯데온_부가세.xlsx", "롯데온에서 다운로드한 파일"),
        ("토스쇼핑_부가세.xlsx", "토스쇼핑에서 다운로드한 파일"),
        ("올웨이즈_부가세.xlsx", "올웨이즈에서 다운로드한 파일"),
    ]

    row += 1
    for col, h in enumerate(["#", "파일명", "설명"], 1):
        ws.cell(row=row, column=col, value=h)
    style_header(ws, row, 3)

    for i, (fname, desc) in enumerate(file_rules, 1):
        row += 1
        for col, val in enumerate([i, fname, desc], 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_cell(cell, align="center" if col == 1 else "left")

    # 추가 확인 사항
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="⚠️ 대표 추가 확인 항목")
    ws.cell(row=row, column=1).font = Font(name="맑은 고딕", bold=True, size=11, color="CC0000")

    extra_items = [
        "세금계산서 없는 현금거래 매출 확인",
        "세금계산서 없는 현금거래 매입 확인",
        "고정자산 매입 (장비, 차량 등) 해당시",
        "수정세금계산서 발행 내역 확인",
    ]
    for item in extra_items:
        row += 1
        ws.cell(row=row, column=2, value=f"• {item}")
        ws.cell(row=row, column=2).font = Font(name="맑은 고딕", size=9)
        ws.merge_cells(f"B{row}:F{row}")

    # ── 셀러센터 URL 시트 ──
    ws2 = wb.create_sheet("셀러센터 URL")
    ws2.column_dimensions['A'].width = 4
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 50

    ws2.cell(row=1, column=1, value="셀러센터 바로가기")
    ws2.cell(row=1, column=1).font = Font(name="맑은 고딕", bold=True, size=12, color=NAVY)
    ws2.merge_cells("A1:C1")

    for col, h in enumerate(["#", "플랫폼", "URL"], 1):
        ws2.cell(row=3, column=col, value=h)
    style_header(ws2, 3, 3)

    for i, p in enumerate(PLATFORMS, 1):
        row2 = i + 3
        for col, val in enumerate([i, p["name"], p["seller_url"]], 1):
            cell = ws2.cell(row=row2, column=col, value=val)
            style_cell(cell, align="center" if col == 1 else "left")
        # URL을 하이퍼링크로
        ws2.cell(row=row2, column=3).hyperlink = p["seller_url"]
        ws2.cell(row=row2, column=3).font = Font(name="맑은 고딕", color="0563C1", underline="single", size=10)

    # 저장
    output_path = OUTPUT_DIR / f"부가세_체크리스트_{quarter}.xlsx"
    wb.save(output_path)
    return output_path


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 법인세 체크리스트 생성
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def create_corp_checklist(year):
    wb = Workbook()
    ws = wb.active
    ws.title = "법인세 체크리스트"

    widths = [4, 40, 20, 15, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:E1")
    cfg = load_config()
    ws['A1'] = f"{cfg['company_name']} 법인세 신고 자료 체크리스트 ({year}년)"
    ws['A1'].font = Font(name="맑은 고딕", bold=True, size=14, color=NAVY)

    ws.merge_cells("A2:E2")
    ws['A2'] = f"신고기한: {int(year)+1}.03.31  |  세무사: {cfg['accountant_name']}  |  작성일: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(name="맑은 고딕", size=9, color="666666")

    row = 4
    headers = ["#", "항목", "출처", "담당", "완료"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header(ws, row, len(headers))

    num = 1
    for category, items in CORP_CHECKLIST:
        row += 1
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws.cell(row=row, column=1, value=category)
        cell.font = Font(name="맑은 고딕", bold=True, size=10, color=NAVY)
        cell.fill = PatternFill(start_color=LIGHT_NAVY, end_color=LIGHT_NAVY, fill_type="solid")

        for item, source, responsible in items:
            row += 1
            values = [num, item, source, responsible, "☐"]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                style_cell(cell, align="center" if col in [1, 5] else "left")
            if "⚠️" in responsible:
                ws.cell(row=row, column=4).font = Font(name="맑은 고딕", color="CC0000", size=10)
            num += 1

    output_path = OUTPUT_DIR / f"법인세_체크리스트_{year}.xlsx"
    wb.save(output_path)
    return output_path


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 카톡 메시지 생성
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def create_kakao_message(tax_type, period):
    cfg = load_config()
    if tax_type == "vat":
        # input 폴더에서 수집된 파일 확인
        input_subdir = INPUT_DIR / period
        collected = []
        if input_subdir.exists():
            for f in input_subdir.iterdir():
                if f.suffix in ['.xlsx', '.xls', '.pdf', '.csv']:
                    collected.append(f.name)

        lines = [
            f"[{cfg['company_name']}] {period} 부가세 자료",
            "=" * 30,
            "",
        ]

        if collected:
            lines.append(f"■ 수집 완료: {len(collected)}개 플랫폼")
            for f in sorted(collected):
                lines.append(f"  ✅ {f}")
            missing = [p["name"] for p in PLATFORMS
                       if not any(p["filename"] in f for f in collected)]
            if missing:
                lines.append(f"\n■ 미수집: {len(missing)}개")
                for m in missing:
                    lines.append(f"  ❌ {m}")
        else:
            lines.append("■ 플랫폼별 부가세 자료 첨부")
            for p in PLATFORMS:
                lines.append(f"  • {p['name']}")

        lines.extend([
            "",
            "■ 추가 확인사항",
            "  • 세금계산서 없는 현금거래: 없음",
            "  • 고정자산 매입: 없음",
            "  • 수정세금계산서: 없음",
            "  (해당사항 있으면 수정해주세요)",
            "",
            "체크리스트 파일 같이 보냅니다.",
            "",
            "━" * 30,
            f"{cfg['company_name']} {cfg['representative']}",
        ])
    else:
        lines = [
            f"[{cfg['company_name']}] {period}년 법인세 자료",
            "=" * 30,
            "",
            "■ 첨부 자료",
            "  1. 이카운트 재무제표",
            "  2. 연간 매출/매입장",
            "  3. 법인 통장 거래내역",
            "",
            "■ 확인사항",
            "  • 감가상각 자산: 없음",
            "  • 가수금/가지급금: 없음",
            "  (해당사항 있으면 수정해주세요)",
            "",
            "체크리스트 파일 같이 보냅니다.",
            "",
            "━" * 30,
            f"{cfg['company_name']} {cfg['representative']}",
        ]

    text = "\n".join(lines)
    label = "부가세" if tax_type == "vat" else "법인세"
    output_path = OUTPUT_DIR / f"{label}_카톡메시지_{period}.txt"
    output_path.write_text(text, encoding="utf-8")
    return output_path, text


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# input 폴더 파일 수집 현황 체크
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def check_collected_files(quarter):
    """수집된 파일 현황 확인"""
    input_subdir = INPUT_DIR / quarter
    input_subdir.mkdir(exist_ok=True)

    print(f"\n📂 파일 수집 현황 ({input_subdir}/)")
    print("─" * 45)

    collected = 0
    for p in PLATFORMS:
        files = [f for f in input_subdir.iterdir()
                 if f.suffix in ['.xlsx', '.xls', '.pdf', '.csv']
                 and p["filename"] in f.name] if input_subdir.exists() else []
        if files:
            print(f"  ✅ {p['name']:12s} → {files[0].name}")
            collected += 1
        else:
            print(f"  ☐  {p['name']:12s} → 미수집")

    print("─" * 45)
    print(f"  {collected}/{len(PLATFORMS)} 플랫폼 수집 완료")
    return collected


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 메인
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    parser = argparse.ArgumentParser(description="세무 자료 패키징")
    parser.add_argument("--type", choices=["vat", "corp"], default="vat")
    parser.add_argument("--quarter", default=None, help="분기 (예: 2026Q1)")
    parser.add_argument("--year", default=None, help="연도 (예: 2025)")
    args = parser.parse_args()

    now = datetime.now()

    if args.type == "vat":
        q = (now.month - 1) // 3 + 1
        period = args.quarter or f"{now.year}Q{q}"
        label = "부가세"

        cfg = load_config()
        print("=" * 50)
        print(f"  {cfg['company_name']} {label} 자료 수집 & 패키징")
        print(f"  기간: {period}")
        print("=" * 50)

        # 체크리스트 생성
        checklist_path = create_vat_checklist(period)
        print(f"\n📋 체크리스트: {checklist_path}")

        # 수집 현황
        check_collected_files(period)

        # 카톡 메시지
        kakao_path, kakao_text = create_kakao_message("vat", period)
        print(f"\n💬 카톡 메시지: {kakao_path}")

    else:
        period = args.year or str(now.year - 1)
        label = "법인세"

        cfg = load_config()
        print("=" * 50)
        print(f"  {cfg['company_name']} {label} 자료 패키징")
        print(f"  기간: {period}년")
        print("=" * 50)

        checklist_path = create_corp_checklist(period)
        print(f"\n📋 체크리스트: {checklist_path}")

        kakao_path, kakao_text = create_kakao_message("corp", period)
        print(f"\n💬 카톡 메시지: {kakao_path}")

    print(f"\n{'─'*50}")
    print("카톡 전달용 메시지:")
    print(f"{'─'*50}")
    print(kakao_text)
    print(f"{'─'*50}")
    print(f"\n✅ 체크리스트 + 플랫폼 자료 파일을 세무사에게 카톡으로 전달하세요.")


if __name__ == "__main__":
    main()
